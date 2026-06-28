"""Excel -> Opportunity ingestion via LLM column mapping.

Two stages:
  1. Show ONLY the column headers to an LLM (OpenRouter) -> it returns a mapping
     of each spreadsheet column to one of our Opportunity fields.
  2. Extract every row deterministically in code using that mapping.

The LLM only interprets header names (cheap, one small call). The actual cell
values are read by code — never sent to the LLM — so nothing is hallucinated and
it scales to large files. Unmapped columns are preserved in `extra`.
"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

import httpx
from openpyxl import load_workbook

from app.settings import settings
from models.opportunity import Opportunity

# Our target fields (name -> hint), used to guide the LLM's column mapping.
_SCHEMA_FIELDS: dict[str, str] = {
    "title": "opportunity title",
    "solicitation_number": "solicitation or RFP number",
    "notice_id": "SAM.gov notice id",
    "agency": "department / agency / office",
    "naics": "NAICS code",
    "psc_code": "PSC / classification code",
    "set_aside": "set-aside type (WOSB, 8(a), SDVOSB, HUBZone, Small Business, Full & Open)",
    "opp_type": "notice type (Solicitation, Sources Sought, Presolicitation, ...)",
    "posted_date": "posted date",
    "response_deadline": "response due date",
    "estimated_value": "estimated value in USD",
    "place_of_performance": "place of performance",
    "poc_name": "point of contact name",
    "poc_email": "point of contact email",
    "description": "short description / notes",
    "link": "source URL",
    "document_url": "link/URL or path to the full solicitation PDF / PWS / SOW document",
}


def _read_sheet(path: str | Path) -> tuple[list[str], list[tuple]]:
    """Return (headers, data_rows) from the active sheet."""
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = ["" if h is None else str(h).strip() for h in rows[0]]
    return headers, rows[1:]


def _map_columns_via_llm(headers: list[str], sample_rows: list[tuple]) -> dict[str, str | None]:
    """Ask the LLM to map each spreadsheet header to one of our fields (or null).

    A few sample values per column are included so ambiguous headers (e.g.
    "Number") can be disambiguated by what's actually under them.
    """
    fields = "\n".join(f"- {name}: {hint}" for name, hint in _SCHEMA_FIELDS.items())

    # show each header with a few example values to guide the mapping
    preview_lines = []
    for idx, header in enumerate(headers):
        if not header:
            continue
        examples = []
        for row in sample_rows:
            val = row[idx] if idx < len(row) else None
            if val is not None:
                examples.append(str(val)[:60])
        preview_lines.append(f'- "{header}": {examples}')
    preview = "\n".join(preview_lines)

    prompt = (
        "Map each spreadsheet column header to one of our opportunity fields. "
        "Each column is shown with a few example values to help you decide.\n\n"
        f"Our fields:\n{fields}\n\n"
        f"Spreadsheet columns (header + sample values):\n{preview}\n\n"
        'Return ONLY JSON of the form: {"mapping": {"<exact column header>": '
        '"<our field name, or null>"}}.\n'
        "Map a header to null if none of our fields fit. Use each of our fields "
        "at most once."
    )
    resp = httpx.post(
        f"{settings.OPENROUTER_BASE_URL}/chat/completions",
        headers={
            "Authorization": f"Bearer {settings.OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        },
        json={
            "model": settings.EXTRACTION_MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "response_format": {"type": "json_object"},
            "temperature": 0,
        },
        timeout=60.0,
    )
    resp.raise_for_status()
    data = json.loads(resp.json()["choices"][0]["message"]["content"])
    return data.get("mapping", {}) if isinstance(data, dict) else {}


def _parse_value(raw) -> float | None:
    """Parse '$22,000,000', '7,500,000', '$12.3M', '1.2B' -> float."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().lower()
    mult = 1.0
    if s.endswith("k"):
        mult = 1e3
    elif s.endswith("m"):
        mult = 1e6
    elif s.endswith("b"):
        mult = 1e9
    num = re.sub(r"[^0-9.]", "", s)
    try:
        return float(num) * mult if num else None
    except ValueError:
        return None


def _stringify(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    return s or None


def ingest_excel(path: str | Path) -> list[Opportunity]:
    """Read an .xlsx file and return normalized Opportunity records.

    The LLM maps the columns once; rows are then extracted deterministically.
    """
    headers, rows = _read_sheet(path)
    if not headers:
        return []

    mapping = _map_columns_via_llm(headers, rows[:3])
    valid_fields = set(Opportunity.model_fields)

    # column index -> our field (or None), keeping the original header for `extra`
    col_field: list[tuple[int, str | None, str]] = []
    for idx, header in enumerate(headers):
        mapped_field = mapping.get(header)
        col_field.append((idx, mapped_field if mapped_field in valid_fields else None, header))

    opportunities: list[Opportunity] = []
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        mapped: dict = {}
        extra: dict = {}
        for idx, field, header in col_field:
            val = row[idx] if idx < len(row) else None
            if field is None:
                if val is not None:
                    extra[header] = _stringify(val)
                continue
            mapped[field] = _parse_value(val) if field == "estimated_value" else _stringify(val)

        if not mapped.get("title"):
            if not mapped.get("solicitation_number"):
                continue
            mapped["title"] = mapped["solicitation_number"]

        mapped["source"] = "excel"
        mapped["extra"] = extra
        try:
            opportunities.append(Opportunity(**mapped))
        except Exception:
            continue

    return opportunities
